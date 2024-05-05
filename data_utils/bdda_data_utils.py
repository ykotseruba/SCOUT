# BDDA dataset utilities

import os
import time
#import getch
import pickle as pkl
import pandas as pd
import numpy as np
import openpyxl
import plotly.express as px
import plotly.io as pio
#pio.kaleido.scope.mathjax = None
from matplotlib_venn import venn2, venn2_circles
from matplotlib import pyplot as plt
from os.path import join, abspath, isfile, isdir, splitext
from os import makedirs, listdir
import cv2
import json
from tqdm import tqdm 
import fire
from tabulate import tabulate

extra_annot_path = os.path.join(os.environ['EXTRA_ANNOT_PATH'], 'BDD-A')
dataset_path = os.environ['BDDA_PATH']

class BDDAUtils():
	def __init__(self, cached=True):

		with open(f'{extra_annot_path}/exclude_videos.json', 'r') as fid:
			exclude_videos = json.load(fid)

		self.video_ids = {} # map from data type to video
		self.vid2type = {} # map from video to data type
		for data_type in ['training', 'validation', 'test']:
			self.video_ids[data_type] = sorted(list(set([int(x) for x in os.listdir(os.path.join(dataset_path,data_type,'camera_frames'))]) 
											- set([x for x in exclude_videos[data_type]])))
			for vid_id in self.video_ids[data_type]:
				self.vid2type[int(vid_id)] = data_type

		self._valid_frames_df = self.get_valid_gt_frames()
		self._vehicle_df = None
		self.load_vehicle_data(cached=cached)


	def print_dataset_stats(self):
		"""
		number of videos
		number of frames
		number of valid videos
		number of valid frames
		number of eye-tracking data points
		"""

		num_videos = 0
		num_frames = 0
		
		video_lengths = []
		for data_type in ['test', 'training', 'validation']:
			video_dirs = os.listdir(os.path.join(dataset_path,data_type,'camera_frames'))
			video_lens = [len(os.listdir(os.path.join(dataset_path, data_type, 'camera_frames', x))) for x in video_dirs]
			num_videos += len(video_dirs)
			num_frames += sum(video_lens)
			video_lengths.extend(video_lens)

		video_lengths = np.array(video_lengths)
		vid_length_mean = np.mean(video_lengths)
		vid_length_std = np.std(video_lengths)

		num_valid_videos = len(self._vehicle_df['vid_id'].drop_duplicates())
		num_valid_frames = len(self._vehicle_df)


		#num_subjects_per_video = self._gaze_df[['subj_id', 'vid_id']].drop_duplicates()

		valid_vid_length = self._vehicle_df.groupby(by='vid_id').count()['frame_id']
		valid_vid_length_mean = valid_vid_length.mean()
		valid_vid_length_std = valid_vid_length.std()


		table = [['# videos', num_videos],
				 ['# frames', num_frames],
				 ['Video length (frames)', f'{vid_length_mean:0.2f}({vid_length_std:0.2f})'],
				 ['# valid videos', num_valid_videos],
				 ['# valid frames', num_valid_frames],
				 ['Valid video length (frames)', f'{valid_vid_length_mean:0.2f}({valid_vid_length_std:0.2f})'],
				]
		print(tabulate(table))


	def get_video_attributes(self):
		vid_attr_file = join(extra_annot_path, 'video_labels.xlsx')
		video_attributes = pd.read_excel(vid_attr_file)
		return video_attributes

	def print_video_issue_stats(self):
		vid_attr = self.get_video_attributes()

		vids_w_issues = vid_attr.dropna(subset=['issues'])
		tot_vids_w_issues = len(vids_w_issues)
		tot_vids = len(vid_attr)

		print(f'Videos with issues: {tot_vids_w_issues} ({tot_vids_w_issues/tot_vids*100:02f})')

		print(vids_w_issues.groupby(by=['type']).count()['vid_id'])

		vid_issues = []
		for idx, row in vid_attr.iterrows():
			issues = row['issues']
			if not pd.isna(issues):
				#print(issues)
				issues = [x.split('(')[0] for x in issues.split(';')]
				vid_issues.extend([{'type': row['type'], 'vid_id': row['vid_id'], 'issues': x.strip()} for x in issues])
			else:
				vid_issues.append({'type': row['type'], 'vid_id': row['vid_id'], 'issues': issues})

		vid_issues = pd.DataFrame.from_dict(vid_issues)
		counts_total = vid_issues.groupby(by=['issues']).count().sort_values(['type', 'vid_id'], ascending=False)
		print(counts_total.to_string())
		counts_per_type = vid_issues.groupby(by=['type', 'issues']).count().sort_values(['type', 'vid_id'], ascending=False)
		print(counts_per_type.to_string())


	def load_vehicle_data(self, cached=True):
		print('-> Loading vehicle data...')
		cached_file = 'cache/bdda_vehicle_data.pkl'
		if cached and os.path.exists(cached_file):
			with open(cached_file, 'rb') as fid:
				self._vehicle_df = pkl.load(fid)
		else:
			for data_type, vid_ids in self.video_ids.items():
				for vid_id in tqdm(vid_ids, desc=data_type):
					data_path = f'{extra_annot_path}/vehicle_data/{vid_id}.xlsx'
					#print(f'Loading {data_path}...')
					if not os.path.exists(data_path):
						continue

					veh_df = pd.read_excel(data_path).rename(columns={'frame_id': 'frame_id'})
					if not veh_df['speed'].eq(-1).all(axis=0):
						veh_df['vid_id'] = vid_id
						veh_df['data_type'] = data_type
						if any(['Unnamed' in x for x in veh_df.columns]):
							veh_df = veh_df[veh_df.filter(regex='^(?!Unnamed)').columns]
						if self._vehicle_df is None:
							self._vehicle_df = veh_df
						else:
							self._vehicle_df = pd.concat([self._vehicle_df, veh_df], ignore_index=True)
						#self._vehicle_df_list.append(pd.read_excel(data_path))
					# else:
					# 	os.system(f'libreoffice {data_path}')
			self._vehicle_df = self._vehicle_df.rename(columns={'frame': 'frame_id'})
			with open(cached_file, 'wb') as fid:
				pkl.dump(self._vehicle_df, fid)

	def compute_mean_frame(self, video_type='image', old_annot=False):
		"""Compute mean image for each video in the dataset
			image, gt
		
		Args:
			video_type (str): image (scene view) or gt (saliency)
		"""
		print(f'Computing mean {video_type}...')

		if video_type == 'gt':
			mean_frame_path = join(extra_annot_path, 'mean_frames', f'bdd-a_mean_{video_type}.png')
			video_mean_path_template = join(extra_annot_path, 'mean_frames', f'bdd-a_mean_gt')
			subdir = 'gazemap_frames'
		else:
			mean_frame_path = join(extra_annot_path, 'mean_frames', f'bdd-a_mean_{video_type}.png')
			video_mean_path_template = join(extra_annot_path, 'mean_frames', f'bdd-a_mean_{video_type}')
			subdir = 'camera_frames'


		if os.path.exists(mean_frame_path):
			return cv2.imread(mean_frame_path, cv2.IMREAD_GRAYSCALE if video_type == 'gt' else cv2.IMREAD_COLOR)

		mean_frame = None
		num_vids = 0
		for data_type, video_ids in self.video_ids.items():
			for vid_id in tqdm(video_ids, desc=data_type):
				img_dir = join(dataset_path, data_type, subdir, str(vid_id))

				video_mean_path = f'{video_mean_path_template}_{vid_id}.png'
				if os.path.exists(video_mean_path):
					video_mean_img = cv2.imread(video_mean_path)
				else:
					video_mean_img = None
					frames = [join(img_dir, x) for x in listdir(img_dir)]
					for frame in frames:
						try:
							if video_mean_img is None:
								video_mean_img = cv2.imread(frame).astype(float)
							else:
								video_mean_img += cv2.imread(frame).astype(float)
						except:
							print(f'Error: could not load {frame}!')
							return  
					
					video_mean_img /= len(frames)

					cv2.imwrite(video_mean_path, video_mean_img.astype(int))
					#print(f'Saved mean img to {video_mean_path}...')

				if mean_frame is None:
					mean_frame = video_mean_img.astype(float)
				else:
					mean_frame += video_mean_img.astype(float)
			num_vids  += len(video_ids)
		mean_frame /= num_vids

		cv2.imwrite(mean_frame_path, mean_frame.astype(int))
		print(f'Saved mean img to {mean_frame_path}')

		return mean_frame


	def get_valid_gt_frames(self):

		print('-> Getting valid gt frames...')
		valid_frames_df = []
		for data_type, video_ids in self.video_ids.items():
			for vid_id in tqdm(video_ids, desc=data_type):
				img_dir = join(dataset_path, data_type, 'gazemap_frames', str(vid_id))
				frames = listdir(img_dir)
				valid_frames_df.extend([{'data_type': data_type, 'vid_id': vid_id, 'frame_id': x} for x in range(len(frames))])
		return pd.DataFrame.from_dict(valid_frames_df)

	def get_evaluation_frames(self, obs_length=16, gt=None):
		'''
			Get frames to compute evaluation metrics on 
		'''
		eval_frames_df = self._valid_frames_df.copy(deep=True) 
		eval_frames_df = eval_frames_df[(eval_frames_df['data_type'] == 'test') & (eval_frames_df['frame_id'] > obs_length)]
		tot_frames = len(eval_frames_df)
		no_uturn_frames_df = self._vehicle_df[(self._vehicle_df['lat action'] != 'U-turn') & (self._vehicle_df['lat action'] != 'reverse')][['vid_id', 'frame_id']] # exclude frames with U-turns in them
		#TODO remove reverse frames
		eval_frames_df = eval_frames_df.merge(no_uturn_frames_df, how='inner', left_on=['vid_id', 'frame_id'], right_on=['vid_id', 'frame_id'])[['vid_id', 'frame_id']]
		print('Excluded', tot_frames-len(eval_frames_df), 'frames')
		return eval_frames_df

	def get_eval_dict(self, cache_path=None):
		"""
		Load cached evaluation dict to continue or create a new one 
		"""

		if cache_path is None:
			raise ValueError('ERROR: provide cache path!')

		if os.path.exists(cache_path):
			with open(cache_path, 'rb') as fid:
				eval_dict = pkl.load(fid)
			print(f'Loaded eval dict from {cache_path}')
		else:
			eval_dict = {}

			for vid in self.video_ids['test']:
				eval_dict[vid] = []
				for frame_id in range(len(os.listdir(os.path.join(dataset_path, 'test', 'camera_frames', str(vid))))):
					eval_dict[vid].append({'vid_id': vid,
									  'frame_id': frame_id,
									  #'pred_img_path': pred_img_path,
									  #'pred_img_mtime': pred_img_mtime,
									  #'gt_img_path': gt_img_path,
									  'SIM': None,
									  'IG': None,
									  'AUC': None,
									  'sAUC': None,
									  'KLdiv': None,
									  'KLdiv_1': None,
									  'NSS': None,
									  'CC': None})

			cache_dir = os.path.split(cache_path)[0]
			if not os.path.exists(cache_dir):
				os.makedirs(cache_dir, exist_ok=True)

			with open(cache_path, 'wb') as fid:
				pkl.dump(eval_dict, fid, protocol=pkl.HIGHEST_PROTOCOL)
			print(f'Saved eval dict to {cache_path}')
		return eval_dict

	def get_intersection_frames(self):
		print('-> Getting intersection frames')
		veh_df = self._vehicle_df.copy(deep=True)
		context_df = veh_df[['vid_id', 'frame_id', 'context']].dropna()
		# get context start frame before the intersection
		context_df[['intersection', 'priority']] = context_df['context'].copy().str.split(';', n=2, expand=True)
		context_df.rename(columns={'frame_id':'end_frame'}, inplace=True)

		intersection_df = []
		for index, row in context_df.iterrows():
			end_frame = int(row['end_frame'])
			start_frame = max(0, int(end_frame - 60))
			vid_id = row['vid_id']
			intersection = row['intersection'].strip()
			priority = row['priority'].strip()
			for frame in range(start_frame, end_frame+1):
				intersection_df.append({'vid_id': vid_id, 'frame_id': frame, 'intersection': intersection, 'priority': priority})

		intersection_df = pd.DataFrame.from_dict(intersection_df)
		return context_df, intersection_df


	def label_action_frames(self, veh_df):

		veh_df['lat action'].fillna('drive straight', inplace=True)
		veh_df = veh_df.drop(veh_df[(veh_df['lat action'] == 'U-turn') | (veh_df['lat action'] == 'reverse')].index)
		# https://journals.sagepub.com/doi/pdf/10.3141/2663-17
		# threshold for deceleration event in data is set to -0.4
		bins = [-10000, -0.4, 0.4, 10000]
		labels = ['decelerate', 'maintain', 'accelerate']
		veh_df['lon action'] = pd.cut(self._vehicle_df['acc'], bins=bins, labels=labels).astype(str)
		veh_df.loc[(veh_df['acc'] == 0) & (veh_df['speed'] >= 0) & (veh_df['speed'] < 1), 'lon action'] = 'stopped'	
		veh_df.loc[(veh_df['acc'] == 0) & (veh_df['speed'] >= 0) & (veh_df['speed'] < 1), 'lat action'] = 'stopped'	

		conditions =  [	veh_df['lon action'].eq('stopped') | veh_df['lat action'].eq('stopped'), # stopped
						veh_df['lat action'].eq('drive straight') & veh_df['lon action'].eq('maintain'), # no action
						veh_df['lat action'].ne('drive straight') & veh_df['lon action'].eq('maintain'),  # lat only
						veh_df['lat action'].eq('drive straight') & veh_df['lon action'].eq('decelerate'), # deceleration
						veh_df['lat action'].eq('drive straight') & veh_df['lon action'].eq('accelerate'), # acceleration
						veh_df['lat action'].ne('drive straight') & veh_df['lon action'].ne('maintain') # accelerating after turn is not an action	
						]
		choices = ['stopped', 'maintain', 'lat only', 'dec only', 'acc only', 'lat/lon']
		veh_df['action'] = np.select(conditions, choices, default=None)
		print(len(veh_df), veh_df.groupby(by='action')['action'].count().sum())
		print(veh_df.groupby(by='action')['action'].count()/len(veh_df)*100)
		return veh_df

	def get_action_frames(self, valid_frames_df=None):
		'''
		Get a dataframe with video ids and frame ids for the following actions:
		stopped = vehicle not moving
		no action = driving straight + maintain speed
		lat only = driving straight + decelerating
		lon only = turn left/right or lane change left/right
		lat/lon = intersection of lat only and lon only

		Args:
		action: name of the action
		valid_frames_df: list of valid video and frame ids (e.g. non-empty ground truth files)
		'''

		veh_df = self._vehicle_df.copy(deep=True)
		action_df = self.label_action_frames(veh_df)
		action_df = action_df[(action_df['lat action'] != 'U-turn') & (action_df['lat action'] != 'reverse')] # remove frames with U-turns in them

		if valid_frames_df is not None: # remove empty
			action_df = action_df.merge(valid_frames_df, how='inner', right_on=['vid_id', 'frame_id'], left_on=['vid_id', 'frame_id'])
		return action_df

	def print_driver_action_stats(self):
		# number of actions
		# mean (std) action durations

		# print as a table (df.to_latex())
		# solution https://towardsdatascience.com/pandas-dataframe-group-by-consecutive-same-values-128913875dba
		veh_df = self._vehicle_df.copy(deep=True)
		
		veh_df = self.label_action_frames(veh_df)
		#vid_ids = list(veh_df['vid_id'].unique()) 

		def get_action_stats(action_type, veh_df):
			action_counts = []
			prev_vid = -1
			prev_action = None
			prev_data_type = None
			record = None
			action_record = None

			num_rows = len(veh_df)
			for idx, row in tqdm(veh_df.iterrows(), total=num_rows, desc=action_type):
				cur_vid = row['vid_id']
				cur_action = row[action_type]

				if (cur_action != prev_action) or (cur_vid != prev_vid):
					if record is not None:
						action_counts.append(record)
					record = {'data_type': row['data_type'], 'vid_id': cur_vid, 'action': cur_action, 'start_frame': row['frame_id'], 'end_frame': row['frame_id'], 'num_frames': 1}
				else:
					record['num_frames'] += 1
					record['end_frame'] += 1

				prev_action = cur_action
				prev_vid = cur_vid
				prev_data_type = row['data_type']

			action_counts.append(record)
			
			action_counts_df = pd.DataFrame.from_dict(action_counts)
			temp = action_counts_df[['action', 'num_frames']].groupby('action')
			action_stats_df = pd.concat([temp.count(), temp.sum(), temp.sum()/len(veh_df)*100, temp.mean(), temp.std()], axis=1)
			action_stats_df.columns = ['count', 'sum', 'perc', 'mean', 'std']


			return action_counts_df, action_stats_df


		lat_action_counts_df, lat_action_stats_df = get_action_stats('lat action', veh_df)
		lon_action_counts_df, lon_action_stats_df = get_action_stats('lon action', veh_df)

		print(lon_action_stats_df.round(2).sort_values(by=['perc'], ascending=False).to_string())
		print(lat_action_stats_df.round(2).sort_values(by=['perc'], ascending=False).to_string())

		# print(lon_action_stats_df[['count', 'mean', 'std', 'perc']].round(2).sort_values(by=['perc'], ascending=False).to_latex())
		# print(lat_action_stats_df[['count', 'mean', 'std', 'perc']].round(2).sort_values(by=['perc'], ascending=False).to_latex())
		
		excel_filename = 'bdd-a_task_action.xlsx'
		

		if os.path.exists(excel_filename):
			writer = pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace')
			book = writer.book			
			book.create_sheet('lon_action')
			book.create_sheet('lat_action')
		else:
			writer = pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w')

		lon_action_stats_df.round(2).sort_values(by=['perc'], ascending=False).to_excel(writer, sheet_name='lon_action')
		lat_action_stats_df.round(2).sort_values(by=['perc'], ascending=False).to_excel(writer, sheet_name='lat_action')
		writer.close()

	def print_context_stats(self):
		# context 
		# straight road, curved road, objects present, intersections
		# intersections: signalized/unsignalized, right-of-way/no right-of-way, type of intersection (3-way, 4-way)

		veh_df = self._vehicle_df.copy(deep=True)
		tot_frames = len(veh_df)
		inters_df = self._vehicle_df[['vid_id', 'frame_id', 'context']].dropna()#.groupby(['context'], as_index=False).count()
		inters_df[['inters_type', 'priority']] = inters_df['context'].str.split(';', expand=True)

		inters_stats_df = inters_df.groupby(['inters_type', 'priority']).count()

		excel_filename = 'bdd-a_task_action.xlsx'
		
		if os.path.exists(excel_filename):
			writer = pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace')
			book = writer.book
			book.create_sheet('context')
		else:
			writer = pd.ExcelWriter(excel_filename, engine='openpyxl', mode='w')
		
		inters_stats_df.to_excel(writer, sheet_name='context')
		writer.close()



if __name__ == '__main__':
	ds = BDDAUtils(cached=True)
	fire.Fire(ds)